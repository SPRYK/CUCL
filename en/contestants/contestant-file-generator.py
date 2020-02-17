#-------------------------------------------------------------
#| to rewrite HTML (F{}.html) please edited in template.html |
#-------------------------------------------------------------

#------------------------------------------------------------------
#| to update contestants (dictionary) you should do the following |
#      1. run generateData() in this file                         |
#      2. copy data from collect_data.txt and paste to this file  |
#      3. run generateFile("th") for update TH HTML files         |
#          or generateFile("en") for update EN HTML files         |
#------------------------------------------------------------------

import os;
import openpyxl as xl

n = 30
contestants = {
    'th' : {
        #contestant number 1
        1 : {
            "contest_name" : "F01 POMPAM",
            "nickname" : "ป๋อมแป๋ม",
            "name" : "แทนรัก",
            "surname" : "อัศวเลิศลักษณ์",
            "faculty" : "คณะนิเทศศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 2
        2 : {
            "contest_name" : "F02 NENE",
            "nickname" : "เนเน่",
            "name" : "วิชรดี",
            "surname" : "ไชยวัฒนสกุล",
            "faculty" : "คณะพาณิชยศาสตร์และการบัญชี",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 3
        3 : {
            "contest_name" : "F03 GRACE",
            "nickname" : "เกรซ",
            "name" : "ณัฐณิชา",
            "surname" : "อัมพรตระกูล",
            "faculty" : "คณะนิเทศศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 4
        4 : {
            "contest_name" : "F04 PUNPUN",
            "nickname" : "ปันปัน",
            "name" : "ศนิชตา",
            "surname" : "ศิริจิตจันทร์",
            "faculty" : "คณะศิลปกรรมศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 5
        5 : {
            "contest_name" : "F05 MIMI",
            "nickname" : "มีมี่",
            "name" : "พิชญา",
            "surname" : "บวรปรัส",
            "faculty" : "คณะอักษรศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 6
        6 : {
            "contest_name" : "F06 PRIM",
            "nickname" : "ปริม",
            "name" : "ศรุดา",
            "surname" : "พรหมอินทร์",
            "faculty" : "คณะทันตแพทยศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 7
        7 : {
            "contest_name" : "F07 MILD",
            "nickname" : "มายด์",
            "name" : "สุชัญญา",
            "surname" : "กิตติฤดีกุล",
            "faculty" : "คณะวิทยาศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 8
        8 : {
            "contest_name" : "F08 BAM",
            "nickname" : "แบม",
            "name" : "ภัทราภรณ์",
            "surname" : "โชติวีระสถานนท์",
            "faculty" : "คณะสัตวแพทยศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 9
        9 : {
            "contest_name" : "F09 BEST",
            "nickname" : "เบส",
            "name" : "ภรลภัสสิรี",
            "surname" : "อัครธนศิริโชติ",
            "faculty" : "คณะนิเทศศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 10
        10 : {
            "contest_name" : "F10 WAM",
            "nickname" : "แวม",
            "name" : "ทอฝัน",
            "surname" : "ไวยวุฒิ",
            "faculty" : "คณะศิลปกรรมศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 11
        11 : {
            "contest_name" : "F11 GINNY",
            "nickname" : "จินนี่",
            "name" : "ณัฐณิชา",
            "surname" : "ประทีปนาฏศิริ",
            "faculty" : "คณะพาณิชยศาสตร์และการบัญชี",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 12
        12 : {
            "contest_name" : "F12 PATTY",
            "nickname" : "แพทตี้",
            "name" : "พัทธนันท์",
            "surname" : "กิจเจริญศักดิ์กุล",
            "faculty" : "คณะพาณิชยศาสตร์และการบัญชี",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 13
        13 : {
            "contest_name" : "F13 MINT",
            "nickname" : "มิ้นท์",
            "name" : "ธิฌาน์",
            "surname" : "ธุระชน",
            "faculty" : "คณะอักษรศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 14
        14 : {
            "contest_name" : "F14 JANE",
            "nickname" : "เจน",
            "name" : "พิชญ์สุกานต์",
            "surname" : "จารุโรจน์ปกรณ์",
            "faculty" : "คณะครุศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 15
        15 : {
            "contest_name" : "F15 MIU",
            "nickname" : "หมิว",
            "name" : "ณัชชา",
            "surname" : "เตชะมงคลาภิวัฒน์",
            "faculty" : "สถาบันนวัตกรรมบูรณาการ",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 16
        16 : {
            "contest_name" : "F16 DREAM",
            "nickname" : "ดรีม",
            "name" : "อังคณา",
            "surname" : "ศิรีสุวรรณ",
            "faculty" : "คณะวิทยาศาสตร์การกีฬา",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 17
        17 : {
            "contest_name" : "M01 WOODY",
            "nickname" : "วูดดี้",
            "name" : "ณัฐวุฒิ",
            "surname" : "ชัยบูรณพันธ์กุล",
            "faculty" : "คณะวิศวกรรมศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 18
        18 : {
            "contest_name" : "M02 YUHOO",
            "nickname" : "ยู้ฮู",
            "name" : "อิสยา",
            "surname" : "กิจพัฒนาศิลป์",
            "faculty" : "คณะวิทยาศาสตร์",
            "year" : "ชั้นปีที่ 3"
        },
        #contestant number 19
        19 : {
            "contest_name" : "M03 DRAGON",
            "nickname" : "ดราก้อน",
            "name" : "วีรินทร์",
            "surname" : "พันนา",
            "faculty" : "คณะวิศวกรรมศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 20
        20 : {
            "contest_name" : "M04 JACKY",
            "nickname" : "แจ็คกี้",
            "name" : "ตงเฉิง",
            "surname" : "โจว",
            "faculty" : "คณะอักษรศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 21
        21 : {
            "contest_name" : "M05 BOM",
            "nickname" : "บอม",
            "name" : "ญาณภัทร",
            "surname" : "ปิยนิจดำรงค์",
            "faculty" : "คณะพาณิชยศาสตร์และการบัญชี",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 22
        22 : {
            "contest_name" : "M06 KIM",
            "nickname" : "คิม",
            "name" : "ปัณณธร",
            "surname" : "จิรศาสตร์",
            "faculty" : "คณะเศรษฐศาสตร์",
            "year" : "ชั้นปีที่ 3"
        },
        #contestant number 23
        23 : {
            "contest_name" : "M07 SAFEFY",
            "nickname" : "เซฟฟี่",
            "name" : "จิรวัฒน์",
            "surname" : "เจนไตรชาญ",
            "faculty" : "คณะวิศวกรรมศาสตร์",
            "year" : "ชั้นปีที่ 3"
        },
        #contestant number 24
        24 : {
            "contest_name" : "M08 GUSJAI",
            "nickname" : "กัสจั๋ย",
            "name" : "ภูธดา",
            "surname" : "โพธิ์ตระการ",
            "faculty" : "คณะแพทยศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 25
        25 : {
            "contest_name" : "M09 BOND",
            "nickname" : "บอน",
            "name" : "ธนภัทร",
            "surname" : "ลิมปอารยะกุล",
            "faculty" : "คณะเศรษฐศาสตร์",
            "year" : "ชั้นปีที่ 3"
        },
        #contestant number 26
        26 : {
            "contest_name" : "M10 GUS",
            "nickname" : "กัส",
            "name" : "สาริน",
            "surname" : "ส่งพิริยะกิจ",
            "faculty" : "คณะทันตแพทยศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 27
        27 : {
            "contest_name" : "M11 ICE",
            "nickname" : "ไอซ์",
            "name" : "พีรณัฐ",
            "surname" : "สุขพณิชนันท์",
            "faculty" : "คณะนิเทศศาสตร์",
            "year" : "ชั้นปีที่ 4"
        },
        #contestant number 28
        28 : {
            "contest_name" : "M12 NET",
            "nickname" : "เน็ต",
            "name" : "กสานติ์",
            "surname" : "ไพบูลย์ผล",
            "faculty" : "คณะแพทยศาสตร์",
            "year" : "ชั้นปีที่ 2"
        },
        #contestant number 29
        29 : {
            "contest_name" : "M13 PEAK",
            "nickname" : "พีค",
            "name" : "ณัฐชนน",
            "surname" : "ไชยพงศ์ผาติ",
            "faculty" : "คณะพาณิชยศาสตร์และการบัญชี",
            "year" : "ชั้นปีที่ 1"
        },
        #contestant number 30
        30 : {
            "contest_name" : "M14 PAEAN",
            "nickname" : "พีน",
            "name" : "ณัชพัฒน์",
            "surname" : "ไชยพงศ์ผาติ",
            "faculty" : "คณะเศรษฐศาสตร์",
            "year" : "ชั้นปีที่ 1"
        },
    },
    'en' : {
        #contestant number 1
        1 : {
            "contest_name" : "F01 POMPAM",
            "nickname" : "POMPAM",
            "name" : "Thaenrak",
            "surname" : "Asavalertlak",
            "faculty" : "Faculty of Communication Arts",
            "year" : "1st year"
        },
        #contestant number 2
        2 : {
            "contest_name" : "F02 NENE",
            "nickname" : "NENE",
            "name" : "Wicharadee",
            "surname" : "Chaiwattanasakul",
            "faculty" : "Faculty of Commerce and Accountancy",
            "year" : "1st year"
        },
        #contestant number 3
        3 : {
            "contest_name" : "F03 GRACE",
            "nickname" : "GRACE",
            "name" : "Natnicha",
            "surname" : "Amporntrakul",
            "faculty" : "Faculty of Communication Arts",
            "year" : "2nd year"
        },
        #contestant number 4
        4 : {
            "contest_name" : "F04 PUNPUN",
            "nickname" : "PUNPUN",
            "name" : "Sanichta",
            "surname" : "Sirichitchant",
            "faculty" : "Faculty of Applied Arts",
            "year" : "1st year"
        },
        #contestant number 5
        5 : {
            "contest_name" : "F05 MIMI",
            "nickname" : "MIMI",
            "name" : "Pitchaya",
            "surname" : "Bovornprus",
            "faculty" : "Faculty of Arts",
            "year" : "4th year"
        },
        #contestant number 6
        6 : {
            "contest_name" : "F06 PRIM",
            "nickname" : "PRIM",
            "name" : "Saruda",
            "surname" : "Prom-in",
            "faculty" : "Faculty of Dentistry",
            "year" : "2nd year"
        },
        #contestant number 7
        7 : {
            "contest_name" : "F07 MILD",
            "nickname" : "MILD",
            "name" : "Suchanya",
            "surname" : "Kittirudeekul",
            "faculty" : "Faculty of Science",
            "year" : "4th year"
        },
        #contestant number 8
        8 : {
            "contest_name" : "F08 BAM",
            "nickname" : "BAM",
            "name" : "Pattaporn",
            "surname" : "Chotveerasatanont",
            "faculty" : "Faculty of Veterinary Science",
            "year" : "1st year"
        },
        #contestant number 9
        9 : {
            "contest_name" : "F09 BEST",
            "nickname" : "BEST",
            "name" : "Pornrapatsiri",
            "surname" : "Akarathanasirichot",
            "faculty" : "Faculty of Communication Arts",
            "year" : "4th year"
        },
        #contestant number 10
        10 : {
            "contest_name" : "F10 WAM",
            "nickname" : "WAM",
            "name" : "Torfun",
            "surname" : "Vaiyawuth",
            "faculty" : "Faculty of Applied Arts",
            "year" : "1st year"
        },
        #contestant number 11
        11 : {
            "contest_name" : "F11 GINNY",
            "nickname" : "GINNY",
            "name" : "Natnicha",
            "surname" : "Pratipnatsiri",
            "faculty" : "Faculty of Commerce and Accountancy",
            "year" : "1st year"
        },
        #contestant number 12
        12 : {
            "contest_name" : "F12 PATTY",
            "nickname" : "PATTY",
            "name" : "Pattanan",
            "surname" : "Kitcharoensakkul",
            "faculty" : "Faculty of Commerce and Accountancy",
            "year" : "4th year"
        },
        #contestant number 13
        13 : {
            "contest_name" : "F13 MINT",
            "nickname" : "MINT",
            "name" : "Thishar",
            "surname" : "Thurachon",
            "faculty" : "Faculty of Arts",
            "year" : "2nd year"
        },
        #contestant number 14
        14 : {
            "contest_name" : "F14 JANE",
            "nickname" : "JANE",
            "name" : "Phitsukarn",
            "surname" : "Jarurojprakorn",
            "faculty" : "Faculty of Education",
            "year" : "2nd year"
        },
        #contestant number 15
        15 : {
            "contest_name" : "F15 MIU",
            "nickname" : "MIU",
            "name" : "Natsha",
            "surname" : "Taechamongkalapiwat",
            "faculty" : "Bachelor of Arts and Science in Integrated Innovation",
            "year" : "1st year"
        },
        #contestant number 16
        16 : {
            "contest_name" : "F16 DREAM",
            "nickname" : "DREAM",
            "name" : "Angkhana",
            "surname" : "Srisuwan",
            "faculty" : "Faculty of Sports Science",
            "year" : "2nd year"
        },
        #contestant number 17
        17 : {
            "contest_name" : "M01 WOODY",
            "nickname" : "WOODY",
            "name" : "Nuttawut",
            "surname" : "Chaibooranapankul",
            "faculty" : "Faculty of Engineering",
            "year" : "4th year"
        },
        #contestant number 18
        18 : {
            "contest_name" : "M02 YUHOO",
            "nickname" : "YUHOO",
            "name" : "Isaya",
            "surname" : "Kijpatanasilp",
            "faculty" : "Faculty of Science",
            "year" : "3rd year"
        },
        #contestant number 19
        19 : {
            "contest_name" : "M03 DRAGON",
            "nickname" : "DRAGON",
            "name" : "Weerin",
            "surname" : "Phanna",
            "faculty" : "Faculty of Engineering",
            "year" : "2nd year"
        },
        #contestant number 20
        20 : {
            "contest_name" : "M04 JACKY",
            "nickname" : "JACKY",
            "name" : "DongCheng",
            "surname" : "Zhou",
            "faculty" : "Faculty of Arts",
            "year" : "4th year"
        },
        #contestant number 21
        21 : {
            "contest_name" : "M05 BOM",
            "nickname" : "BOM",
            "name" : "Yanapat",
            "surname" : "Piyanijdamrong",
            "faculty" : "Faculty of Commerce and Accountancy",
            "year" : "2nd year"
        },
        #contestant number 22
        22 : {
            "contest_name" : "M06 KIM",
            "nickname" : "KIM",
            "name" : "Pannathorn",
            "surname" : "Jirasart",
            "faculty" : "Faculty of Economics",
            "year" : "3rd year"
        },
        #contestant number 23
        23 : {
            "contest_name" : "M07 SAFEFY",
            "nickname" : "SAFEFY",
            "name" : "Jirawat",
            "surname" : "Jentricharn",
            "faculty" : "Faculty of Engineering",
            "year" : "3rd year"
        },
        #contestant number 24
        24 : {
            "contest_name" : "M08 GUSJAI",
            "nickname" : "GUSJAI",
            "name" : "Phutada",
            "surname" : "Photrakarn",
            "faculty" : "Faculty of Medicine",
            "year" : "1st year"
        },
        #contestant number 25
        25 : {
            "contest_name" : "M09 BOND",
            "nickname" : "BOND",
            "name" : "Tanaphat",
            "surname" : "Limpaarayakul",
            "faculty" : "Faculty of Economics",
            "year" : "3rd year"
        },
        #contestant number 26
        26 : {
            "contest_name" : "M10 GUS",
            "nickname" : "GUS",
            "name" : "Sarin",
            "surname" : "Songpiriyakij",
            "faculty" : "Faculty of Dentistry",
            "year" : "1st year"
        },
        #contestant number 27
        27 : {
            "contest_name" : "M11 ICE",
            "nickname" : "ICE",
            "name" : "Peeranut",
            "surname" : "Sukpanichnant",
            "faculty" : "Faculty of Communication Arts",
            "year" : "4th year"
        },
        #contestant number 28
        28 : {
            "contest_name" : "M12 NET",
            "nickname" : "NET",
            "name" : "Kasarn",
            "surname" : "Paiboonpol",
            "faculty" : "Faculty of Medicine",
            "year" : "2nd year"
        },
        #contestant number 29
        29 : {
            "contest_name" : "M13 PEAK",
            "nickname" : "PEAK",
            "name" : "Natchanon",
            "surname" : "Chaipongpati",
            "faculty" : "Faculty of Commerce and Accountancy",
            "year" : "1st year"
        },
        #contestant number 30
        30 : {
            "contest_name" : "M14 PAEAN",
            "nickname" : "PAEAN",
            "name" : "Natchapat",
            "surname" : "Chaipongpati",
            "faculty" : "Faculty of Economics",
            "year" : "1st year"
        },
    }
}


# read Sheet Excel to create data in dictionary form
# to generate contestants data in dictionary 
# after generate ditionary you should copy collect_data.txt to this file
def generateData(lang):
    file = lang + '/contestants/contestants_sheet.xlsx'
    wb = xl.load_workbook(file)
    sh = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    code = ""
    code += ("contestants = {") + "\n"
    #--------------------------
    #data_th
    code += (" "*4 + "'th' : {") + "\n"
    for i in range(1,n+1):
        code += (" "*8 + "#contestant number %d" % i) + "\n"
        code += (" "*8 + "%d : {" % i) + "\n"
        code += (" "*12 + '"contest_name" : "{}",'.format(sh.cell(row=i+1,column=1).value)) + "\n"
        code += (" "*12 + '"nickname" : "{}",'.format(sh.cell(row=i+1,column=2).value)) + "\n"
        code += (" "*12 + '"name" : "{}",'.format(sh.cell(row=i+1,column=3).value)) + "\n"
        code += (" "*12 + '"surname" : "{}",'.format(sh.cell(row=i+1,column=4).value)) + "\n"
        code += (" "*12 + '"faculty" : "{}",'.format(sh.cell(row=i+1,column=5).value)) + "\n"
        code += (" "*12 + '"year" : "{}"'.format(sh.cell(row=i+1,column=6).value)) + "\n"
        code += (" "*8 + "},") + "\n"
    code += (" "*4 + "},") + "\n"

    #data_en
    code += (" "*4 + "'en' : {") + "\n"
    for i in range(1,n+1):
        code += (" "*8 + "#contestant number %d" % i) + "\n"
        code += (" "*8 + "%d : {" % i) + "\n"
        code += (" "*12 + '"contest_name" : "{}",'.format(sh.cell(row=i+2+n,column=1).value)) + "\n"
        code += (" "*12 + '"nickname" : "{}",'.format(sh.cell(row=i+2+n,column=2).value)) + "\n"
        code += (" "*12 + '"name" : "{}",'.format(sh.cell(row=i+2+n,column=3).value)) + "\n"
        code += (" "*12 + '"surname" : "{}",'.format(sh.cell(row=i+2+n,column=4).value)) + "\n"
        code += (" "*12 + '"faculty" : "{}",'.format(sh.cell(row=i+2+n,column=5).value)) + "\n"
        code += (" "*12 + '"year" : "{}"'.format(sh.cell(row=i+2+n,column=6).value)) + "\n"
        code += (" "*8 + "},") + "\n"
    code += (" "*4 + "}") + "\n"
    #--------------------------
    code += ("}") + "\n"
    print(code)
    # contain data in collect_data.txt
    f = open(lang + "/contestants/collect_data.txt","w")
    f.write(code)
    f.close()

# to delete all contestants .html file
def deleteAllHtmlFile(lang):
    num = [16,14]
    cid = ["F","M"]
    print("until delete")
    for i in range(len(cid)):
        for j in range(num[i]):
            s = str(j+1) if j+1 >= 10 else "0" + str(j+1)
            contest = cid[i] + s
            fname = "en/contestants/" + contest + ".html"
            if os.path.exists(fname):
                os.remove(fname)
                print(fname,"was deleted")
            else:
                print(fname,"does not exist")

# to delete specific content in file 
def deleteContent(fName):
    with open(fName, "w"):
        pass

# to overwrite specific .html file with contestants data in ditionary 
def genreateHtml(html,index,lang):
    DATA = contestants["th"][index] if lang == "th" else contestants["en"][index]
    html = html.format(DATA["contest_name"].split()[0],
    DATA["contest_name"],
    DATA["nickname"],
    DATA["name"],
    DATA["surname"],
    DATA["faculty"],
    DATA["year"])
    return html

# to generate All .html file
# for ver.th please call generateFile("th")
# for ver.en please call generateFile("en")
def generateFile(lang):
    contents = ""
    print("until genFile")
    t = open(lang + "/contestants/template.html","r");
    if t.mode == "r":
        contents = t.read()
    DATA = contestants["th"] if lang == "th" else contestants["en"]
    for i in range(1,n+1):
        fname = lang + "/contestants/" + DATA[i]["contest_name"].split()[0] + ".html"
        deleteContent(fname)
        f = open(fname,"w")
        f.write(genreateHtml(contents,i,lang))
        f.close()

if __name__ == "__main__":
    print("input lang (th/en) : ",end="")
    lang = input()
    print("input cmd (gendata=0,genfile=1,delfile=2) : ",end="") 
    cmd = int(input())
    if cmd == 0:
        generateData(lang)
    elif cmd == 1:
        generateFile(lang)
    elif cmd == 2:
        deleteAllHtmlFile(lang)
    print("done")
